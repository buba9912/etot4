import os
import subprocess
import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


def new_xlsx(xlsx_title):
    date = time.strftime('%Y-%m-%d_%H-%M', time.localtime())
    xlsx_file_path = "t4_readable_file\\" + xlsx_title + date + ".xlsx"
    # Create a xlsx Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = xlsx_title
    wb.save(xlsx_file_path)
    # open the xlsx
    load_workbook(xlsx_file_path)
    return xlsx_file_path


def new_row_top(text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion, text_Questions, text_Text,
                text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight, text_Assessment_team_info,
                text_Response_team_info, df):
    new_row = pd.DataFrame({'Questionnaire': [text_Questionnaire],
                            'Description': [text_Description],
                            'Criteria': [text_Criteria],
                            'Parent Criterion': [text_Parent_Criterion],
                            'Questions': [text_Questions],
                            'Text': [text_Text],
                            'Legends': [text_Legends],
                            'Possible Answers': [text_Possible_Answers],
                            'P.A Legend': [text_PA_Legend],
                            'Type': [text_Type],
                            'Weight': [text_Weight],
                            'Assessment team info': [text_Assessment_team_info],
                            'Response team info': [text_Response_team_info]})

    df = pd.concat([new_row, df]).reset_index(drop=True)
    return df


def insert_row(idx, df, text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion, text_Questions,
               text_Text,
               text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight, text_Assessment_team_info,
               text_Response_team_info):
    new_row = pd.DataFrame({'Questionnaire': [text_Questionnaire],
                            'Description': [text_Description],
                            'Criteria': [text_Criteria],
                            'Parent Criterion': [text_Parent_Criterion],
                            'Questions': [text_Questions],
                            'Text': [text_Text],
                            'Legends': [text_Legends],
                            'Possible Answers': [text_Possible_Answers],
                            'P.A Legend': [text_PA_Legend],
                            'Type': [text_Type],
                            'Weight': [text_Weight],
                            'Assessment team info': [text_Assessment_team_info],
                            'Response team info': [text_Response_team_info]})

    dfA = df.iloc[:idx, ]
    dfB = df.iloc[idx:, ]

    df = dfA.append(new_row).append(dfB).reset_index(drop=True)

    return df


def refactor_file(df, answer_type):
    # delete unnecessary columns
    df.drop('Unnamed: 7', axis=1, inplace=True)
    df.drop('n: notwendig', axis=1, inplace=True)

    # rename columns
    df = df.rename(columns={'Prüfpunkte Titel (n)': "Questions"})
    df = df.rename(columns={'Prüfpunkt Frage/Anforderung (n)': "Text"})
    df = df.rename(columns={'Unterkriterium (o)': "Criteria"})
    df = df.rename(columns={'Kriterium (n)': "Parent Criterion"})
    df = df.rename(columns={'Information Auditor, z.B. Anforderung aus NORM (o)': "Assessment team info"})
    df = df.rename(columns={'Information Auditee (o), z.B. Ziel': "Response team info"})
    # df = df.rename(columns={'': ""})

    # sort columns in correct order
    df = df.reindex(columns=['Questionnaire',
                             'Description',
                             'Criteria',
                             'Parent Criterion',
                             'Questions',
                             'Text',
                             'Legends',
                             'Possible Answers',
                             'P.A Legend',
                             'Type',
                             'Weight',
                             'Assessment team info',
                             'Response team info',
                             'Norm-Ref. (o)'
                             ])

    # set missing parameters
    last_Parent_Criterion = ""
    for i in range(1, len(df)):
        #
        if pd.isnull(df.at[i, 'Parent Criterion']) is False:
            last_Parent_Criterion = df.at[i, 'Parent Criterion']

        if pd.isnull(df.at[i, 'Criteria']) is False:
            if pd.isnull(df.at[i, 'Parent Criterion']) is True:
                df.loc[[i], 'Parent Criterion'] = last_Parent_Criterion

        # set Parent Criterion as Criteria if Criteria not exist
        if pd.isnull(df.at[i, 'Criteria']) is True:
            df.loc[[i], 'Criteria'] = df.at[i, 'Parent Criterion']
        # set default Weight
        if pd.isnull(df.at[i, 'Questions']) is False:
            if pd.isnull(df.at[i, 'Weight']) is True:
                df.loc[[i], 'Weight'] = 1
        # set ANSWER_TYPE
        if pd.isnull(df.at[i, 'Questions']) is False:
            if pd.isnull(df.at[i, 'Possible Answers']) is True:
                df.loc[[i], 'Possible Answers'] = "ANSWER_TYPE_" + str(answer_type)

            # Norm-Ref. (o) to info
        if pd.isnull(df.at[i, 'Norm-Ref. (o)']) is False:
            df.loc[[i], 'Assessment team info'] = "Norm-Ref.:" + df.at[i, 'Norm-Ref. (o)']
            df.loc[[i], 'Response team info'] = "Norm-Ref.:" + df.at[i, 'Norm-Ref. (o)']

    # delete Norm-Ref. (o) columns
    df.drop('Norm-Ref. (o)', axis=1, inplace=True)

    return df


def make_headline_criteria(df):
    last_Parent_Criterion = ""

    for i in range(1, len(df)):
        if pd.isnull(df.at[i, "Criteria"]) is False:
            if last_Parent_Criterion != df.at[i, 'Parent Criterion']:
                last_Parent_Criterion = df.at[i, 'Parent Criterion']
                # print(last_Parent_Criterion)

                text_Questionnaire = ""
                text_Description = ""
                text_Criteria = df.at[i, 'Parent Criterion']
                text_Parent_Criterion = ""
                text_Questions = ""
                text_Text = ""
                text_Legends = ""
                text_Possible_Answers = ""
                text_PA_Legend = ""
                text_Type = ""
                text_Weight = ""
                text_Assessment_team_info = ""
                text_Response_team_info = ""
                df = insert_row(i, df, text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion,
                                text_Questions,
                                text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                                text_Assessment_team_info, text_Response_team_info)
    # delete unnecessary rows
    df.drop(index=0, axis=0, inplace=True)
    return df


def delit(df):
    pass
    last_last_Parent_Criterion = ""
    last_Parent_Criterion = ""
    for i in range(1, len(df)):
        if pd.isnull(df.at[i, "Parent Criterion"]) is False:
            if last_Parent_Criterion != df.at[i, 'Parent Criterion']:
                last_last_Parent_Criterion = last_Parent_Criterion
                last_Parent_Criterion = df.at[i, 'Parent Criterion']
                if last_last_Parent_Criterion != df.at[i, 'Parent Criterion']:
                    df.at[i, 'Parent Criterion'] = ""
    return df


def shift_questions(df):
    text_Questionnaire = ""
    text_Description = ""
    text_Criteria = ""
    text_Parent_Criterion = ""
    text_Questions = ""
    text_Text = ""
    text_Legends = ""
    text_Possible_Answers = ""
    text_PA_Legend = ""
    text_Type = ""
    text_Weight = ""
    text_Assessment_team_info = ""
    text_Response_team_info = ""

    df = insert_row(len(df) + 1, df, text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion,
                    text_Questions,
                    text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                    text_Assessment_team_info, text_Response_team_info)
    df.Questions = df.Questions.shift(1)
    df.Text = df.Text.shift(1)
    df['Possible Answers'] = df['Possible Answers'].shift(1)
    df['Weight'] = df['Weight'].shift(1)
    df['Assessment team info'] = df['Assessment team info'].shift(1)
    df['Response team info'] = df['Response team info'].shift(1)

    return df


def define_answer_rows(answer_type, df):
    if answer_type == 1:
        text_Questionnaire = ""
        text_Description = ""
        text_Criteria = ""
        text_Parent_Criterion = ""
        text_Questions = ""
        text_Text = ""
        text_Legends = ""
        text_Possible_Answers = "n.a."
        text_PA_Legend = "nicht anwendbar"
        text_Type = "IMP"
        text_Weight = ""
        text_Assessment_team_info = "nicht anwendbar"
        text_Response_team_info = ""

        df = insert_row(3, df, text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion,
                        text_Questions,
                        text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                        text_Assessment_team_info, text_Response_team_info)

        text_Questionnaire = ""
        text_Description = ""
        text_Criteria = ""
        text_Parent_Criterion = ""
        text_Questions = ""
        text_Text = ""
        text_Legends = ""
        text_Possible_Answers = "nicht erfüllt"
        text_PA_Legend = "0% - 15%"
        text_Type = "NOR"
        text_Weight = 100
        text_Assessment_team_info = "Es gibt wenig bis keinen Nachweis, dass ein Attribut eines Prozesses erfüllt wird. (0% - 15%)"
        text_Response_team_info = "Es gibt wenig bis keinen Nachweis, dass ein Attribut eines Prozesses erfüllt wird. (0% - 15%)"
        df = insert_row(3, df, text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion,
                        text_Questions,
                        text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                        text_Assessment_team_info, text_Response_team_info)

        text_Questionnaire = ""
        text_Description = ""
        text_Criteria = ""
        text_Parent_Criterion = ""
        text_Questions = ""
        text_Text = ""
        text_Legends = ""
        text_Possible_Answers = "teilweise erfüllt"
        text_PA_Legend = "16% - 50%"
        text_Type = "NOR"
        text_Weight = 67
        text_Assessment_team_info = "Es gibt einen Nachweis, dass ein Attribut eines Prozesses teilweise erfüllt wird. (15% - 50%)"
        text_Response_team_info = "Es gibt einen Nachweis, dass ein Attribut eines Prozesses teilweise erfüllt wird. (15% - 50%)"
        df = insert_row(3, df, text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion,
                        text_Questions,
                        text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                        text_Assessment_team_info, text_Response_team_info)

        text_Questionnaire = ""
        text_Description = ""
        text_Criteria = ""
        text_Parent_Criterion = ""
        text_Questions = ""
        text_Text = ""
        text_Legends = ""
        text_Possible_Answers = "weitgehend erfüllt"
        text_PA_Legend = "51% - 85%"
        text_Type = "NOR"
        text_Weight = 32
        text_Assessment_team_info = "Es gibt einen Nachweis eines systematischen Ansatzes und des signifikanten Erfüllens eines Attributs eines Prozesses. Es können Schwächen bzgl. des Attributs vorliegen. (50% - 85%"
        text_Response_team_info = "Es gibt einen Nachweis eines systematischen Ansatzes und des signifikanten Erfüllens eines Attributs eines Prozesses. Es können Schwächen bzgl. des Attributs vorliegen. (50% - 85%"
        df = insert_row(3, df, text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion,
                        text_Questions,
                        text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                        text_Assessment_team_info, text_Response_team_info)

        text_Questionnaire = ""
        text_Description = ""
        text_Criteria = ""
        text_Parent_Criterion = ""
        text_Questions = ""
        text_Text = ""
        text_Legends = ""
        text_Possible_Answers = "vollständig erfüllt"
        text_PA_Legend = "86% - 100%"
        text_Type = "NOR"
        text_Weight = 0
        text_Assessment_team_info = "Es gibt einen Nachweis eines vollständigen, systematischen Ansatzes und vollständig erfüllten Attributs eines Prozesses. Keine nennenswerten Schwächen bzgl. des Attributs liegen vor. (85% - 100%)"
        text_Response_team_info = "Es gibt einen Nachweis eines vollständigen, systematischen Ansatzes und vollständig erfüllten Attributs eines Prozesses. Keine nennenswerten Schwächen bzgl. des Attributs liegen vor. (85% - 100%)"
        df = insert_row(3, df, text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion,
                        text_Questions,
                        text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                        text_Assessment_team_info, text_Response_team_info)

    return df


# Add 3 rows on top
def define_damage_and_name_rows(df):
    text_Questionnaire = "Mittel"
    text_Description = 25000
    text_Criteria = ""
    text_Parent_Criterion = ""
    text_Questions = ""
    text_Text = ""
    text_Legends = ""
    text_Possible_Answers = ""
    text_PA_Legend = ""
    text_Type = ""
    text_Weight = ""
    text_Assessment_team_info = ""
    text_Response_team_info = ""
    df = new_row_top(text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion, text_Questions,
                     text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                     text_Assessment_team_info, text_Response_team_info, df)

    text_Questionnaire = "Hoch"
    text_Description = 50000
    text_Criteria = ""
    text_Parent_Criterion = ""
    text_Questions = ""
    text_Text = ""
    text_Legends = ""
    text_Possible_Answers = ""
    text_PA_Legend = ""
    text_Type = ""
    text_Weight = ""
    text_Assessment_team_info = ""
    text_Response_team_info = ""
    df = new_row_top(text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion, text_Questions,
                     text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                     text_Assessment_team_info, text_Response_team_info, df)

    text_Questionnaire = "Gering"
    text_Description = 5000
    text_Criteria = ""
    text_Parent_Criterion = ""
    text_Questions = ""
    text_Text = ""
    text_Legends = ""
    text_Possible_Answers = ""
    text_PA_Legend = ""
    text_Type = ""
    text_Weight = ""
    text_Assessment_team_info = ""
    text_Response_team_info = ""
    df = new_row_top(text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion, text_Questions,
                     text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                     text_Assessment_team_info, text_Response_team_info, df)

    text_Questionnaire = "!AUSFÜLLEN"
    text_Description = "Beschreibung Vorlage Assessment"
    text_Criteria = ""
    text_Parent_Criterion = ""
    text_Questions = ""
    text_Text = ""
    text_Legends = ""
    text_Possible_Answers = ""
    text_PA_Legend = ""
    text_Type = ""
    text_Weight = ""
    text_Assessment_team_info = "# Anleitung und Erklärungen zum vorliegenden Assessment"
    text_Response_team_info = "# Anleitung und Erklärungen zum vorliegenden Assessment"
    df = new_row_top(text_Questionnaire, text_Description, text_Criteria, text_Parent_Criterion, text_Questions,
                     text_Text, text_Legends, text_Possible_Answers, text_PA_Legend, text_Type, text_Weight,
                     text_Assessment_team_info, text_Response_team_info, df)
    return df


def generate_file(df, xlsx_file_fame):
    # print(df)
    cols = list(df.columns.values)
    # print(cols)
    df.to_excel(xlsx_file_fame, index=False, header=True)


def open_file_and_folder(xlsx_file_path):
    # open file
    os.system(xlsx_file_path)

    # open explorer
    p = Path(xlsx_file_path).resolve()
    for i in range(0, len(str(p))):
        if str(p)[-1] != "\\":
            p = str(p)[: -1]
        else:
            p = str(p)[: -1]
            break
    print(p)

    path = os.path.normpath(p)

    if sys.platform == "linux" or sys.platform == "linux2":
        print(p)
    elif sys.platform == "darwin":
        print(p)
        pass
    # OS X
    elif sys.platform == "win32":
        FILEBROWSER_PATH = os.path.join(os.getenv('WINDIR'), 'explorer.exe')
        if os.path.isdir(path):
            subprocess.run([FILEBROWSER_PATH, path])
        elif os.path.isfile(path):
            subprocess.run([FILEBROWSER_PATH, '/select,', os.path.normpath(path)])


# Windows...

# subprocess.Popen(r'p')


def convert_to_t4_excel(xlsx_title, answer_type, input_filename):
    # open file
    df: object = pd.read_excel(input_filename)

    xlsx_file_path = new_xlsx(xlsx_title)

    df = refactor_file(df, answer_type)

    df = make_headline_criteria(df)

    # df = delit(df)

    df = shift_questions(df)

    df = define_answer_rows(answer_type, df)

    df = define_damage_and_name_rows(df)

    generate_file(df, xlsx_file_path)

    open_file_and_folder(xlsx_file_path)


def main():
    xlsx_title = "main"
    answer_type = 1
    input_filename = "C:\\Users\\Mika F\\Documents\\etot4\\Template_Fragebogen_K4_Design.xlsx"

    convert_to_t4_excel(xlsx_title, answer_type, input_filename)


if __name__ == '__main__':
    main()
